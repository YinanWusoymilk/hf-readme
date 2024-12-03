import os
import markdown
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

def extract_headers_and_content_sequential(file_path):
    """Extract headers and their corresponding content in the order they appear."""
    with open(file_path, 'r', encoding='utf-8') as f:
        md_content = f.read()

    # Convert Markdown to HTML
    html = markdown.markdown(md_content)

    # Parse the HTML using BeautifulSoup
    soup = BeautifulSoup(html, 'html.parser')

    # Store the headers and their following content sequentially
    headers_content = []
    for element in soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6']):
        header_level = element.name  # Capture header level (e.g., 'h1')
        header_text = element.text.strip()

        # Extract the content following the header
        content = []
        next_sibling = element.find_next_sibling()
        while next_sibling and next_sibling.name not in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            content.append(next_sibling.text.strip())
            next_sibling = next_sibling.find_next_sibling()

        # Combine header and content with the header name bolded
        full_text = f"**{header_text} ({header_level.upper()})**\n" + "\n".join(content)
        headers_content.append(full_text)

    return headers_content

def process_folder(folder_path, output_file):
    """Process all Markdown files in the folder and write results to Excel."""
    data = []

    # Iterate through all Markdown files in the folder
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.md'):
                file_path = os.path.join(root, file)
                headers_content = extract_headers_and_content_sequential(file_path)

                # Prepare a row where each header-content pair is in a unique column
                row = {'Filename': file}
                for idx, header_content in enumerate(headers_content):
                    # Store the header-content pair and an empty comment column
                    row[f'Header {idx + 1}'] = header_content
                    row[f'Comment {idx + 1}'] = ""  # Empty comment column
                data.append(row)

    # Create a DataFrame and convert it to Excel with formatting
    df = pd.DataFrame(data)
    max_headers = max((len(row) - 1) // 2 for row in data)  # Max number of headers

    # Reorder columns for consistent structure
    column_order = ['Filename'] + [
        f'{label} {i + 1}' for i in range(max_headers) for label in ['Header', 'Comment']
    ]
    df = df.reindex(columns=column_order, fill_value='')

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write the DataFrame to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        if r_idx > 0:  # Skip header row
            for c_idx, cell_value in enumerate(row[1:], start=2):  # Skip 'Filename' column
                if isinstance(cell_value, str) and 'Header' in ws.cell(1, c_idx).value:
                    # Apply bold formatting to header part and wrap text
                    header_text, content_text = (
                        cell_value.split('\n', 1) if '\n' in cell_value else (cell_value, '')
                    )
                    cell_obj = ws.cell(row=r_idx + 1, column=c_idx)
                    cell_obj.value = f"{header_text}\n{content_text}"
                    cell_obj.alignment = Alignment(wrap_text=True)

    # Save the workbook
    wb.save(output_file)

# Example usage
folder_path = '../readme_files/'
output_file = 'markdown_headers_with_comments.xlsx'
process_folder(folder_path, output_file)

print(f'Headers and content extracted and saved to {output_file}.')
